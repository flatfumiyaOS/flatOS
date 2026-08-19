"""
工程表(xlsx)生成スクリプト。

「【例】工程表.xlsx」のスタイル(パステルカラー、列幅、日付/曜日フォーマットなど)を
踏襲したガントチャート形式の工程表を作る。表紙エリア(1〜5行目)は_write_cover()が
自動で書き込む。

使い方:
    python3 build_schedule.py <config.json> <output.xlsx>

config.json のスキーマ:
{
  "title": "大蔵様邸　改修工事",   # 現状はセルには書き込まないが、呼び出し側の管理用に残す
  "start_date": "2026-07-13",   # 着工日 (YYYY-MM-DD)。日付の表示開始(B列)はこの日を含む週の月曜日になる
  "total_days": 21,              # 表に含める実働日数(工期の目安から決める)
  "sheet_name": "工程表",         # 省略可
  "rows": [
    {
      "label": "仮設工事",        # A列に入る工種名
      "noise": false,             # true にすると警告色(赤)で塗る。近隣配慮工事の行で使う
      "tasks": [
        {"start_day": 1, "end_day": 2, "text": "養生"}
        # start_day/end_day は着工日を1日目とした実働日の通し番号。1日だけなら start_day==end_day
      ]
    }
  ]
}

同じ行の tasks は日数が重ならないようにすること(同じ列に2つの内容は入れられない)。
別の行同士なら重なってよい(工事の並行作業)。

同じ工種で複数の作業がある場合(例: 造作工事の中にボード下地組み・ボード貼りがある)は、
rows に同じ label で複数のエントリを連続して並べること。連続する同じ label は自動的に
A列が縦結合され、同じ色で塗られる(1エントリ=1行、結合はA列のみで、マス自体は結合しない)。

日付列は工期がどんなに短くてもAT列まで(MIN_CALENDAR_COLUMNS参照)自動的に延長される。
"""
import sys
import json
import datetime
from pathlib import Path

import jpholiday
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 「【例】工程表.xlsx」から抽出したGoogleスプレッドシート標準パステルカラー
COLORS = ['FFF4CCCC', 'FFFFF2CC', 'FFD9EAD3', 'FFD9D2E9', 'FFC9DAF8', 'FFD9D9D9', 'FFEAD1DC']
NOISE_COLOR = 'FFE06666'

BASE_FONT = Font(name='Arial', size=10)
TITLE_FONT = Font(name='Arial', size=14, bold=True)
LABEL_FONT = Font(name='Arial', size=10, bold=True)
TASK_FONT = Font(name='Arial', size=10)
NOISE_TASK_FONT = Font(name='Arial', size=10, color='FFFFFF')
DATE_FONT = Font(name='Arial', size=9)
COVER_LABEL_FONT = Font(name='Arial', size=9, bold=True)
COVER_VALUE_FONT = Font(name='Arial', size=9)
COVER_NOTE_FONT = Font(name='Arial', size=8)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_TITLE = Alignment(horizontal='left', vertical='center')
LEFT_TOP_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
# マス(9B〜、実際の工事内容セル)は結合しないため、隣のセルへ見た目上はみ出させる。
# wrap_textをTrueにすると自セル内で折り返してしまいはみ出さなくなるためFalseにする。
TASK_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=False)

GRID_SIDE = Side(style='thin', color='FFD9D9D9')
GRID_BORDER = Border(left=GRID_SIDE, right=GRID_SIDE, top=GRID_SIDE, bottom=GRID_SIDE)

LOGO_PATH = Path(__file__).parent / "assets" / "corp_logo.png"

# 表紙に毎回同じ内容で入る固定値（会社名・担当者など）
COVER_COMPANY_TITLE = '株式会社フラット\n工程表\n\n'
COVER_SITE_SUPERVISOR = '平居 靖弘・平居 史也（株式会社フラット）'
COVER_SALES_REP = '平居 靖弘（株式会社フラット）'
COVER_NOTE = '作業の進捗状況により予告なく変更する場合がございます。\n予めご了承ください。'


WEEKEND_HOLIDAY_FILL = PatternFill(fill_type='solid', fgColor='FFF4CCCC')  # 薄い赤


def _is_working_day(d: datetime.date) -> bool:
    return d.weekday() < 5 and not jpholiday.is_holiday(d)


MIN_CALENDAR_COLUMNS = 45  # B列〜AT列の列数。工期がどんなに短くてもAT列まで日付を出す。


def _build_calendar(start: datetime.date, n_working_days: int):
    """着工日の週の月曜日から、実働日数がn_working_days分になるまでの連続した暦日リストを作る。

    日付の開始(B列)は着工日そのものではなく、着工日を含む週の月曜日にする
    （例: 6/4(木)着工なら6/1(月)から表示）。月曜日〜着工日前日は実働日として
    数えず、単に日付・曜日だけを表示する空欄期間になる。
    土日・祝日も歯抜けにせず列として並べる（作業内容の予定を入れないだけ）。
    「着工日を1日目とした通し日数」で組まれたタスクの日数(start_day/end_day)は
    実働日の通し番号を指すため、実働日の通し番号→暦日リスト内のインデックスの
    対応表も合わせて返す。着工日そのものが土日・祝日の場合は、その次の実働日から
    1日目として数える（工期は常に実働日数として扱うため）。
    工期が短く実働日数分だけではMIN_CALENDAR_COLUMNSに満たない場合は、
    AT列まで日付だけを埋めるために暦日を延長する。
    """
    week_monday = start - datetime.timedelta(days=start.weekday())
    calendar_dates = []
    working_index_to_calendar_index = {}
    working_count = 0
    d = week_monday
    while working_count < n_working_days:
        calendar_dates.append(d)
        if d >= start and _is_working_day(d):
            working_count += 1
            working_index_to_calendar_index[working_count] = len(calendar_dates) - 1
        d += datetime.timedelta(days=1)
    while len(calendar_dates) < MIN_CALENDAR_COLUMNS:
        calendar_dates.append(d)
        d += datetime.timedelta(days=1)
    return calendar_dates, working_index_to_calendar_index


COVER_ROWS = 5  # 表紙用に確保する行数（本体をこの分だけ下にずらす。1〜5行目は空欄のまま）


def _write_cover(ws) -> None:
    """表紙エリア（1〜5行目）に、現場名・現場住所などの入力欄と固定情報を書き込む。

    「現場名」「現場住所」は現場監督が案件ごとに手入力する想定のため、ラベルのみ
    書き込み、値は空欄のままにする。工事担当・営業担当・注意書きは毎回同じ内容の
    固定値として書き込む。
    """
    ws.merge_cells('C1:D2')
    ws['C1'] = '現場名'
    ws['C1'].font = COVER_LABEL_FONT
    ws['C1'].alignment = LEFT_TOP_WRAP

    ws.merge_cells('C3:L4')
    ws['C3'] = '現場住所：'
    ws['C3'].font = COVER_LABEL_FONT
    ws['C3'].alignment = LEFT_TOP_WRAP

    ws.merge_cells('A2:A4')
    ws['A2'] = COVER_COMPANY_TITLE
    ws['A2'].font = TITLE_FONT
    ws['A2'].alignment = LEFT_TOP_WRAP

    ws.merge_cells('N1:W1')
    ws['N1'] = '工事情報'
    ws['N1'].font = COVER_LABEL_FONT
    ws['N1'].alignment = LEFT_TITLE

    ws.merge_cells('N2:O2')
    ws['N2'] = '工事期間'
    ws['N2'].font = COVER_LABEL_FONT
    ws['N2'].alignment = LEFT_TITLE
    ws['P2'] = '〜'
    ws['P2'].font = COVER_VALUE_FONT
    ws['P2'].alignment = LEFT_TITLE

    ws.merge_cells('N3:O3')
    ws['N3'] = '工事担当'
    ws['N3'].font = COVER_LABEL_FONT
    ws['N3'].alignment = LEFT_TITLE
    ws['P3'] = COVER_SITE_SUPERVISOR
    ws['P3'].font = COVER_VALUE_FONT
    ws['P3'].alignment = LEFT_TITLE

    ws.merge_cells('N4:O4')
    ws['N4'] = '営業担当'
    ws['N4'].font = COVER_LABEL_FONT
    ws['N4'].alignment = LEFT_TITLE
    ws['P4'] = COVER_SALES_REP
    ws['P4'].font = COVER_VALUE_FONT
    ws['P4'].alignment = LEFT_TITLE

    ws.merge_cells('Y1:AK1')
    ws['Y1'] = '備考'
    ws['Y1'].font = COVER_LABEL_FONT
    ws['Y1'].alignment = LEFT_TITLE

    ws.merge_cells('Y2:AK4')
    ws['Y2'] = COVER_NOTE
    ws['Y2'].font = COVER_NOTE_FONT
    ws['Y2'].alignment = LEFT_TOP_WRAP

    # ロゴ画像（AM2:AT3セルを結合し、その中に貼り付ける）
    ws.merge_cells('AM2:AT3')
    if LOGO_PATH.exists():
        img = XLImage(str(LOGO_PATH))
        # セル結合の見た目に収まる程度のサイズに縮小する（元画像は960x204）
        img.width = 240
        img.height = 51
        ws.add_image(img, 'AM2')


def build(config: dict, out_path: str):
    start = datetime.date.fromisoformat(config['start_date'])
    n_working_days = int(config['total_days'])
    dates, working_col = _build_calendar(start, n_working_days)
    n_cols = len(dates)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = config.get('sheet_name', '工程表')

    # 列幅: A=工種名(広め), B=着工日(狭め,例に合わせる), C以降=標準幅
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 4.25
    for i in range(3, n_cols + 2):
        ws.column_dimensions[get_column_letter(i)].width = 12.63

    first_col = 2  # B列が1日目

    # 工程表本体は表紙の分だけ下にずらす（元の1〜3行目相当が、ここでは
    # month_row〜weekday_rowになる。列側はA=工種名、B=1日目のまま変更しない）。
    # 1〜5行目（COVER_ROWS分）は空欄のまま残す。
    month_row = 1 + COVER_ROWS
    date_row = 2 + COVER_ROWS
    weekday_row = 3 + COVER_ROWS
    start_row = 4 + COVER_ROWS

    # 表紙エリアの行の高さ: 1行目・6行目(month_row)はデフォルトのまま、
    # 2〜4行目は31、5行目は11にする。
    ws.row_dimensions[2].height = 31
    ws.row_dimensions[3].height = 31
    ws.row_dimensions[4].height = 31
    ws.row_dimensions[5].height = 11

    _write_cover(ws)

    month_ranges = []
    seg_start = 0
    for i in range(1, n_cols + 1):
        if i == n_cols or dates[i].month != dates[seg_start].month:
            month_ranges.append((seg_start, i - 1))
            seg_start = i
    for s, e in month_ranges:
        c1 = get_column_letter(first_col + s)
        c2 = get_column_letter(first_col + e)
        d = dates[s]
        ws.merge_cells(f'{c1}{month_row}:{c2}{month_row}')
        cell = ws[f'{c1}{month_row}']
        cell.value = f'{d.year}年{d.month}月'
        cell.font = BASE_FONT
        cell.alignment = LEFT_TITLE

    # date_row: 日付 (m/d), weekday_row: 曜日 (date_rowを参照する数式、ddd表示)
    ws[f'A{weekday_row}'] = '工程名'
    ws[f'A{weekday_row}'].font = LABEL_FONT
    last_row = weekday_row + len(config['rows'])
    for i, d in enumerate(dates):
        col = get_column_letter(first_col + i)
        c2 = ws[f'{col}{date_row}']
        c2.value = d
        c2.number_format = 'm/d'
        c2.font = DATE_FONT
        c2.alignment = CENTER
        c3 = ws[f'{col}{weekday_row}']
        c3.value = f'={col}{date_row}'
        c3.number_format = 'ddd'
        c3.font = DATE_FONT
        c3.alignment = CENTER
        if not _is_working_day(d):
            # 土日・祝日は列ごと薄い赤で塗って、休みだと分かるようにする
            # （後で工種の行を塗るときに、実際に予定が入っている日は上書きされる）。
            for r in range(date_row, last_row + 1):
                ws.cell(row=r, column=first_col + i).fill = WEEKEND_HOLIDAY_FILL
    ws.row_dimensions[date_row].height = 18
    ws.row_dimensions[weekday_row].height = 18

    # start_row以降: 工種ごとの行。
    # 同じ工種(label)が連続する場合は、A列を縦結合してひとつの分類として扱い、
    # 分類内では同じ色を使う（例: 造作工事の中に複数の工事がある場合）。
    groups = []
    for row in config['rows']:
        if (
            groups
            and groups[-1][0]['label'] == row['label']
            and bool(groups[-1][0].get('noise')) == bool(row.get('noise'))
        ):
            groups[-1].append(row)
        else:
            groups.append([row])

    r = start_row
    for group_idx, group in enumerate(groups):
        group_start_row = r
        is_noise = bool(group[0].get('noise'))
        color = NOISE_COLOR if is_noise else COLORS[group_idx % len(COLORS)]
        fill = PatternFill(fill_type='solid', fgColor=color)
        task_font = NOISE_TASK_FONT if is_noise else TASK_FONT

        for row in group:
            ws.row_dimensions[r].height = 30
            # マス(9B〜)はセルを結合せず、テキストは左揃えで隣のセルにはみ出させる。
            for task in row['tasks']:
                d1, d2, text = task['start_day'], task['end_day'], task['text']
                i1 = working_col[d1]
                i2 = working_col[d2]
                c1 = get_column_letter(first_col + i1)
                cell = ws[f'{c1}{r}']
                cell.value = text
                cell.font = task_font
                cell.alignment = TASK_ALIGN
                for col_idx in range(first_col + i1, first_col + i2 + 1):
                    ws.cell(row=r, column=col_idx).fill = fill
            r += 1

        a = ws[f'A{group_start_row}']
        a.value = group[0]['label']
        a.font = LABEL_FONT
        a.alignment = Alignment(horizontal='left', vertical='center')
        if r - 1 > group_start_row:
            ws.merge_cells(f'A{group_start_row}:A{r - 1}')

    # A6〜（月ヘッダーから最終行まで）に薄いグレーの格子(枠線)を入れる。
    end_col = first_col + n_cols - 1
    for grid_row in range(month_row, last_row + 1):
        for grid_col in range(1, end_col + 1):
            ws.cell(row=grid_row, column=grid_col).border = GRID_BORDER

    ws.freeze_panes = f'B{start_row}'
    wb.save(out_path)
    return out_path


if __name__ == '__main__':
    if len(sys.argv) != 3:
        print('usage: python3 build_schedule.py <config.json> <output.xlsx>')
        sys.exit(1)
    with open(sys.argv[1], encoding='utf-8') as f:
        cfg = json.load(f)
    path = build(cfg, sys.argv[2])
    print('saved:', path)
